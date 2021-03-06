# -*- coding: utf-8 -*-

import openpyxl
import math
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, Style
import base64
import cStringIO
from openerp import models, fields, api, tools, exceptions as e
from datetime import datetime
from calendar import monthrange
import datetime as d
from dateutil import relativedelta

from babel.dates import format_datetime


MONTHS = [(1, 'January'),
          (2, 'February'),
          (3, 'March'),
          (4, 'April'),
          (5, 'May'),
          (6, 'June'),
          (7, 'July'),
          (8, 'August'),
          (9, 'September'),
          (10, 'October'),
          (11, 'November'),
          (12, 'December')]

YEARS = [(2016, '2016'),
         (2017, '2017'),
         (2018, '2018'),
         (2019, '2019'),
         (2020, '2020')]


def print_date(date):
    dayy = datetime.strptime(date, tools.DEFAULT_SERVER_DATE_FORMAT)
    day = format_datetime(dayy, format="EEEE", locale="de")
    res = day[0] + day[1] + format_datetime(dayy, format=", dd.LL.", locale="de")
    return res


def format_float_time_str(time):
    x = math.modf(time)
    decimal = abs(int(x[0] * 61))
    whole = abs(int(x[1]))
    if decimal >= 60:
        decimal -= 60
        whole += 1
    res = str(whole) if whole >= 10 else '0'+str(whole)
    res += ':'
    res += str(decimal) if decimal >= 10 else '0'+str(decimal)
    res += ':00'
    return res


def format_float_time(time):
    x = math.modf(time)
    decimal = abs(int(x[0] * 61))
    whole = abs(int(x[1]))
    if decimal >= 60:
        decimal -= 60
        whole += 1
    return d.time(whole, decimal)


class EmployeeTimesheetGeneratorLine(models.TransientModel):
    _name = 'employee.timesheet.generator.line'

    employee_timesheet_generator_id = fields.Many2one('employee.timesheet.generator', 'Timesheet generator')
    employee_id = fields.Many2one('hr.employee', 'Employee')
    job_id = fields.Many2one('hr.job', 'Job')

    @api.one
    @api.onchange('employee_id')
    def onchange_employee(self):
        if self.employee_id and self.employee_id.job_id:
            self.job_id = self.employee_id.job_id


class EmployeeTimesheetGenerator(models.TransientModel):
    _name = 'employee.timesheet.generator'

    def _reopen(self, res_id):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Generate XLS timesheets',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': res_id,
            'nodestroy': True,
            'view_id': self.env.ref('project_timesheet_activities.employee_timesheet_generator_form'),
            'res_model': 'employee.timesheet.generator',
            'target': 'new',
        }

    @api.one
    @api.depends('employee_timesheet_generator_line_ids.employee_id')
    def _get_lines_count(self):
        if self.employee_timesheet_generator_line_ids:
            self.lines_count = len(self.employee_timesheet_generator_line_ids)

    def _get_default_month(self):
        return datetime.now().month

    def _get_default_year(self):
        return datetime.now().year

    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    month = fields.Selection(MONTHS, 'Month', default=_get_default_month)
    year = fields.Selection(YEARS, 'Year', default=_get_default_year)
    department_id = fields.Many2one('hr.department', 'Department')
    employee_timesheet_generator_line_ids = fields.One2many('employee.timesheet.generator.line',
                                                            'employee_timesheet_generator_id',
                                                            string='Employees')
    state = fields.Selection([('choose', 'Choose'), ('get', 'Get')], 'State', default='choose')
    lines_count = fields.Integer(compute=_get_lines_count, string='Employees count')
    display_timesheets = fields.Boolean('Timesheets', default=False)
    display_sap_report = fields.Boolean('SAP report', default=False)
    use_period = fields.Boolean('Use period', default=True)
    period_id = fields.Many2one('hr.timesheet.sap.period', string='Period')
    sap_date_from = fields.Date('SAP Date from', default=datetime.now()+relativedelta.relativedelta(weekday=0, days=-6))
    sap_date_to = fields.Date('SAP Date to', default=datetime.now()+relativedelta.relativedelta(weekday=6))

    @api.constrains('sap_date_from', 'sap_date_to')
    def sap_dates_constrains(self):
        if self.sap_date_from and self.sap_date_to and self.sap_date_to < self.sap_date_from:
            raise e.ValidationError('SAP REPORT: Date from must be before date to.')

    @api.multi
    def fill_lines(self):
        emp_table = self.env['hr.employee']
        emp_gen_lines = self.env['employee.timesheet.generator.line']
        emp_objs = emp_table.search([('department_id', '=', self.department_id.id), ('active', '=', True)])
        for emp in emp_objs:
            local_dict = {
                'employee_timesheet_generator_id': self.id,
                'employee_id': emp.id,
                'job_id': emp.job_id.id,
            }
            emp_gen_lines.create(local_dict)
        return {
            'type': 'ir.actions.act_window',
            'res_id': self[0].id,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'employee.timesheet.generator',
            'target': 'new',
            'context': self.env.context,
        }

    @api.multi
    def get_report(self):
        this = self[0]
        month = this.month
        year = this.year

        if this.use_period:
            sap_from = this.period_id.period_from
            sap_to = this.period_id.period_to
        else:
            sap_from = this.sap_date_from
            sap_to = this.sap_date_to

        start_date = datetime(year, month, 1)
        end_date = datetime(year, month, monthrange(year, month)[1])

        sheet_lines = self.env['account.analytic.line']
        wb = openpyxl.Workbook(encoding='utf-8')

        left = Border(left=Side(color=colors.BLACK, border_style='thick'))
        right = Border(right=Side(color=colors.BLACK, border_style='thick'))
        top = Border(top=Side(color=colors.BLACK, border_style='thick'))
        bottom = Border(bottom=Side(color=colors.BLACK, border_style='thick'))



        if not self.env.user._has_group(self.env.cr, self.env.user.id, 'base.group_hr_user'):
            employees = self.env['hr.employee'].search([('user_id', '=', self.env.user.id)])
        elif len(this.employee_timesheet_generator_line_ids) == 0:
            employees = self.env['hr.employee'].search([])
        else:
            employees = [item.employee_id for item in this.employee_timesheet_generator_line_ids]

        if this.display_timesheets is True and len(employees) > 0:

            for employee in employees:

                ws = wb.create_sheet(employee.name, 0)

                write_header(ws, employee, month, year)

                iteration_date = start_date
                n = [0]
                days_index = 0
                working_time_sum = 0.0
                working_time_per_day_sum = 0.0
                ws['F8'] = ''
                while iteration_date <= end_date:
                    days_index += 1
                    color = 'e0eaff'
                    if days_index % 2 == 0:
                        color = '9bbcff'

                    day_off = self.env['hr.day.off'].search([('date', '=', iteration_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT))], limit=1)
                    if day_off:
                        color = '999a9e'

                    leave_request = self.env['hr.holidays'].search([('employee_id', '=', employee.id),
                                                                    ('type', '=', 'remove'),
                                                                    ('date_from', '<=', iteration_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)),
                                                                    ('date_to', '>=', iteration_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)),
                                                                    ('state', 'in', ['validate', 'validate1', 'confirm'])],
                                                                   limit=1)

                    # ovdje ispitati da li je day_off ili samo nema linija
                    analytic_lines = sheet_lines.search([('user_id', '=', employee.user_id.id),
                                                         ('timesheet_sheet_id', '!=', False),
                                                         ('date', '=', iteration_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT))],
                                                        order='timesheet_start_time ASC')
                    if len(analytic_lines) < 1:
                        # nema zapisa da je radio, pa treba napraviti jedan prazan red
                        n[0] += 1

                        if day_off and day_off.category_id.name == 'Public Holiday':
                            write_pub_holiday_line(ws, n, color, iteration_date)
                            ws['F'+str(7+n[0])] = format_float_time_str(7.695)
                            ws['F'+str(7+n[0])].style = Style(number_format="HH:MM:SS",
                                                              fill=PatternFill(patternType='solid',
                                                                               fill_type='solid',
                                                                               fgColor=Color(color)),
                                                              border=Border(left=Side(style='thick', color=colors.BLACK),
                                                                            right=Side(style='thick', color=colors.BLACK),
                                                                            top=Side(style='thin', color=colors.BLACK),
                                                                            bottom=Side(style='thin', color=colors.BLACK)))
                            working_time_per_day_sum += 7.695
                            working_time_sum += 7.695

                        elif day_off and day_off.category_id.name == 'Weekend':
                            write_line(ws, n, color, iteration_date)

                        elif leave_request:
                            start_time = leave_request.holiday_status_id.timesheet_start_time if leave_request.holiday_status_id else 8.0
                            end_time = leave_request.holiday_status_id.timesheet_end_time if leave_request.holiday_status_id else 15.695
                            sum_time = abs(end_time - start_time) if start_time and end_time else 0.0
                            write_leave_request_line(ws, n, color, leave_request.holiday_status_id.name, start_time, end_time, sum_time, iteration_date)
                            ws['F'+str(7+n[0])] = format_float_time_str(sum_time)
                            ws['F'+str(7+n[0])].style = Style(number_format="HH:MM:SS",
                                                              fill=PatternFill(patternType='solid',
                                                                               fill_type='solid',
                                                                               fgColor=Color(color)),
                                                              border=Border(left=Side(style='thick', color=colors.BLACK),
                                                                            right=Side(style='thick', color=colors.BLACK),
                                                                            top=Side(style='thin', color=colors.BLACK),
                                                                            bottom=Side(style='thin', color=colors.BLACK)))
                            working_time_per_day_sum += sum_time
                            working_time_sum += sum_time

                        else:
                            write_line(ws, n, color, iteration_date)

                        iteration_date += d.timedelta(days=1)
                        continue
                    else:
                        working_time_per_day = 0.0
                        for index, line in enumerate(analytic_lines):
                            working_time_sum += line.unit_amount
                            working_time_per_day += line.unit_amount
                            n[0] += 1
                            if index == len(analytic_lines)-1:
                                if working_time_per_day != 0:
                                    ws['F'+str(7+n[0])] = format_float_time_str(working_time_per_day)
                                    ws['F'+str(7+n[0])].style = Style(number_format="HH:MM:SS")
                                working_time_per_day_sum += working_time_per_day

                            write_line(ws, n, color, iteration_date, line)
                        iteration_date += d.timedelta(days=1)



                write_footer(ws, n, working_time_sum, working_time_per_day_sum)

        if this.display_sap_report is True and len(employees) > 0:

            if this.use_period:
                unlocked_periods = self.env['hr.timesheet.sap.period'].search([('is_locked', '=', False),
                                                                               ('period_from', '<=', this.period_id.period_from)],
                                                                              order='period_from')
                for current_period in unlocked_periods:
                    lines = sheet_lines.search([('user_id', 'in', [item.user_id.id for item in employees]),
                                                ('timesheet_sheet_id', '!=', False),
                                                # ('task_id', '!=', False),
                                                ('project_activity_id.show_on_sap_report', '=', True),
                                                ('date', '>=', current_period.period_from),
                                                ('date', '<=', current_period.period_to)],
                                               order='user_id, date, write_date')

                    ws = wb.create_sheet(current_period.period_name, 0)

                    ws['A1'] = 'Personalnummer *'
                    ws['A1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['B1'] = 'Datum *'
                    # ws['B1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['C1'] = 'Beginn (HH:MM)'
                    ws['C1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['D1'] = 'Ende (HH:MM)'
                    ws['D1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['E1'] = 'Dauer (HH:MM) *'
                    ws['E1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['F1'] = 'Servicenummer *'
                    ws['F1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['G1'] = 'Zeitart'
                    ws['G1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['H1'] = 'Zuschlag'
                    ws['H1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['I1'] = 'Aufgabennummer *'
                    ws['I1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['J1'] = 'Abweichende abrechenbare Dauer (HH:MM)'
                    ws['J1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['K1'] = 'Arbeitsbeschreibung'
                    ws['K1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['L1'] = 'Field of activity'
                    ws['L1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['M1'] = 'Work Package'
                    ws['M1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['N1'] = 'Interner Kommentar'
                    ws['N1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['O1'] = 'Mitarbeiter'
                    ws['O1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['P1'] = 'Comment'
                    ws['P1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['Q1'] = 'Create date'
                    ws['Q1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                    ws['R1'] = 'Posted on SAP'
                    ws['R1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))

                    ws.row_dimensions[1].height = 50
                    ws.column_dimensions['M'].width = 20
                    ws.column_dimensions['N'].width = 20
                    ws.column_dimensions['O'].width = 20


                    n = 1
                    for line in lines:
                        n += 1
                        line_color = '000000'  # black color

                        emp = self.env['hr.employee'].search([('user_id', '=', line.user_id.id)], limit=1)
                        project_wp_line = self.env['project.activity.work.package.line'].search([('account_id', '=', line.account_id.id),
                                                                                                 ('work_package_id', '=', line.project_activity_work_package_id.id)])

                        ws['A'+str(n)] = emp.other_id if emp and emp.other_id else '---'
                        ws['A'+str(n)].font = Font(color=Color(line_color))
                        ws['B'+str(n)] = d.datetime.strptime(line.date, tools.DEFAULT_SERVER_DATE_FORMAT)
                        ws['B'+str(n)].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'),
                                                     number_format="DD.MM.YYYY")
                        ws['B'+str(n)].font = Font(color=Color(line_color))
                        ws['C'+str(n)] = format_float_time(line.timesheet_start_time)
                        ws['C'+str(n)].font = Font(color=Color(line_color))
                        ws['D'+str(n)] = format_float_time(line.timesheet_end_time)
                        ws['D'+str(n)].font = Font(color=Color(line_color))
                        ws['E'+str(n)] = format_float_time(line.unit_amount)
                        ws['E'+str(n)].font = Font(color=Color(line_color))
                        service_number = '---'
                        if line.project_activity_work_package_id.sap_report_service_number:
                            service_number = line.project_activity_work_package_id.sap_report_service_number
                        if project_wp_line and len(project_wp_line) == 1 and project_wp_line[0].sap_report_service_number:
                            service_number = project_wp_line[0].sap_report_service_number
                        if service_number != '---' and line.project_activity_id.name and 'Travel' in line.project_activity_id.name:
                            service_number += '_R'
                        ws['F'+str(n)] = service_number
                        ws['F'+str(n)].font = Font(color=Color(line_color))
                        ws['G'+str(n)] = ''
                        ws['G'+str(n)].font = Font(color=Color(line_color))
                        ws['H'+str(n)] = ''
                        ws['H'+str(n)].font = Font(color=Color(line_color))

                        task_prefix = ''
                        task_sufix = line.project_activity_work_package_id.sap_report_task_sufix if line.project_activity_work_package_id.sap_report_task_sufix else ''
                        if project_wp_line and len(project_wp_line) == 1:
                            task_prefix = project_wp_line[0].sap_report_task_prefix if project_wp_line[0].sap_report_task_prefix else ''
                            task_sufix = project_wp_line[0].sap_report_task_sufix if project_wp_line[0].sap_report_task_sufix else ''
                        ws['I'+str(n)] = task_prefix + line.task_id.name + task_sufix if line.task_id else ''
                        ws['I'+str(n)].font = Font(color=Color(line_color))

                        ws['J'+str(n)] = ''
                        ws['J'+str(n)].font = Font(color=Color(line_color))
                        ws['K'+str(n)] = ''
                        ws['K'+str(n)].font = Font(color=Color(line_color))
                        ws['L'+str(n)] = line.account_id.name if line.account_id else ''
                        ws['L'+str(n)].font = Font(color=Color(line_color))
                        ws['M'+str(n)] = line.project_activity_work_package_id.name if line.project_activity_work_package_id else ''
                        ws['M'+str(n)].font = Font(color=Color(line_color))
                        ws['N'+str(n)] = line.project_activity_id.name if line.project_activity_id else ''
                        ws['N'+str(n)].font = Font(color=Color(line_color))
                        ws['O'+str(n)] = line.user_id.name
                        ws['O'+str(n)].font = Font(color=Color(line_color))
                        ws['P'+str(n)] = line.timesheet_comment if line.timesheet_comment else ''
                        ws['P'+str(n)].font = Font(color=Color(line_color))
                        ws['Q'+str(n)] = d.datetime.strptime(line.create_date, tools.DEFAULT_SERVER_DATETIME_FORMAT)
                        ws['Q'+str(n)].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'),
                                                     number_format="DD.MM.YYYY")
                        ws['Q'+str(n)].font = Font(color=Color(line_color))
                        line_posted_on_sap = '---'
                        if current_period.last_post:
                            if line.create_date <= current_period.last_post:
                                line_posted_on_sap = 'Yes'
                            else:
                                line_posted_on_sap = 'No'
                                line_color = 'e80000'  # red color
                        else:
                            line_posted_on_sap = 'No'
                            line_color = 'e80000'  # red color
                        ws['R'+str(n)] = line_posted_on_sap
                        ws['R'+str(n)].font = Font(color=Color(line_color))

                    # hr leaves records
                    lv_current_date = d.datetime.strptime(sap_from, tools.DEFAULT_SERVER_DATE_FORMAT)
                    lv_end_date = d.datetime.strptime(sap_to, tools.DEFAULT_SERVER_DATE_FORMAT)
                    while lv_current_date <= lv_end_date:
                        lv_is_holiday = self.env['hr.day.off'].search([('date', '=', lv_current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT))], limit=1)
                        if lv_is_holiday:
                            lv_current_date += d.timedelta(days=1)
                            continue

                        lv_lines = self.env['hr.holidays'].search([('date_from', '<=', lv_current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)),
                                                                   ('date_to', '>=', lv_current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)),
                                                                   ('state', '=', 'validate')])
                        if lv_lines and len(lv_lines)>0:
                            lv_line_color = 'ff8d02'
                            for line in lv_lines:
                                n += 1
                                emp = line.employee_id

                                ws['A'+str(n)] = emp.other_id if emp and emp.other_id else '---'
                                ws['A'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['B'+str(n)] = lv_current_date
                                ws['B'+str(n)].style = Style(font=Font(color=Color(lv_line_color)), alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'),
                                                             number_format="DD.MM.YYYY")

                                ws['C'+str(n)] = format_float_time(8.0)
                                ws['C'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['D'+str(n)] = format_float_time(15.695)
                                ws['D'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['E'+str(n)] = format_float_time(7.695)
                                ws['E'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['F'+str(n)] = '---'
                                ws['F'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['G'+str(n)] = ''
                                ws['G'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['H'+str(n)] = ''
                                ws['H'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['I'+str(n)] = ''
                                ws['I'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['J'+str(n)] = ''
                                ws['J'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['K'+str(n)] = ''
                                ws['K'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['L'+str(n)] = line.holiday_status_id.name
                                ws['L'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['M'+str(n)] = ''
                                ws['M'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['N'+str(n)] = ''
                                ws['N'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['O'+str(n)] = emp.user_id.name if emp.user_id else 'No related user for: ' + emp.name
                                ws['O'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['P'+str(n)] = ''
                                ws['P'+str(n)].font = Font(color=Color(lv_line_color))

                                ws['Q'+str(n)] = d.datetime.strptime(line.create_date, tools.DEFAULT_SERVER_DATETIME_FORMAT)
                                ws['Q'+str(n)].style = Style(font=Font(color=Color(lv_line_color)),
                                                             alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'),
                                                             number_format="DD.MM.YYYY")
                                line_posted_on_sap = '---'
                                ws['R'+str(n)] = line_posted_on_sap
                        lv_current_date += d.timedelta(days=1)

            else:
                lines = sheet_lines.search([('user_id', 'in', [item.user_id.id for item in employees]),
                                            ('timesheet_sheet_id', '!=', False),
                                            # ('task_id', '!=', False),
                                            ('project_activity_id.show_on_sap_report', '=', True),
                                            ('date', '>=', sap_from),
                                            ('date', '<=', sap_to)],
                                           order='user_id, date, write_date')

                ws = wb.create_sheet('SAP REPORT', 0)

                ws['A1'] = 'Personalnummer *'
                ws['A1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['B1'] = 'Datum *'
                # ws['B1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['C1'] = 'Beginn (HH:MM)'
                ws['C1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['D1'] = 'Ende (HH:MM)'
                ws['D1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['E1'] = 'Dauer (HH:MM) *'
                ws['E1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['F1'] = 'Servicenummer *'
                ws['F1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['G1'] = 'Zeitart'
                ws['G1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['H1'] = 'Zuschlag'
                ws['H1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['I1'] = 'Aufgabennummer *'
                ws['I1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['J1'] = 'Abweichende abrechenbare Dauer (HH:MM)'
                ws['J1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['K1'] = 'Arbeitsbeschreibung'
                ws['K1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['L1'] = 'Field of activity'
                ws['L1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['M1'] = 'Work Package'
                ws['M1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['N1'] = 'Interner Kommentar'
                ws['N1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['O1'] = 'Mitarbeiter'
                ws['O1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['P1'] = 'Comment'
                ws['P1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['Q1'] = 'Create date'
                ws['Q1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                ws['R1'] = 'Posted on SAP'
                ws['R1'].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))

                ws.row_dimensions[1].height = 50
                ws.column_dimensions['M'].width = 20
                ws.column_dimensions['N'].width = 20
                ws.column_dimensions['O'].width = 20


                n = 1
                for line in lines:
                    n += 1
                    line_color = '000000'  # black color

                    emp = self.env['hr.employee'].search([('user_id', '=', line.user_id.id)], limit=1)
                    project_wp_line = self.env['project.activity.work.package.line'].search([('account_id', '=', line.account_id.id),
                                                                                             ('work_package_id', '=', line.project_activity_work_package_id.id)])

                    ws['A'+str(n)] = emp.other_id if emp and emp.other_id else '---'
                    ws['A'+str(n)].font = Font(color=Color(line_color))
                    ws['B'+str(n)] = d.datetime.strptime(line.date, tools.DEFAULT_SERVER_DATE_FORMAT)
                    ws['B'+str(n)].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'),
                                                 number_format="DD.MM.YYYY")
                    ws['B'+str(n)].font = Font(color=Color(line_color))
                    ws['C'+str(n)] = format_float_time(line.timesheet_start_time)
                    ws['C'+str(n)].font = Font(color=Color(line_color))
                    ws['D'+str(n)] = format_float_time(line.timesheet_end_time)
                    ws['D'+str(n)].font = Font(color=Color(line_color))
                    ws['E'+str(n)] = format_float_time(line.unit_amount)
                    ws['E'+str(n)].font = Font(color=Color(line_color))
                    service_number = '---'
                    if line.project_activity_work_package_id.sap_report_service_number:
                        service_number = line.project_activity_work_package_id.sap_report_service_number
                    if project_wp_line and len(project_wp_line) == 1 and project_wp_line[0].sap_report_service_number:
                        service_number = project_wp_line[0].sap_report_service_number
                    if service_number != '---' and line.project_activity_id.name and 'Travel' in line.project_activity_id.name:
                        service_number += '_R'
                    ws['F'+str(n)] = service_number
                    ws['F'+str(n)].font = Font(color=Color(line_color))
                    ws['G'+str(n)] = ''
                    ws['G'+str(n)].font = Font(color=Color(line_color))
                    ws['H'+str(n)] = ''
                    ws['H'+str(n)].font = Font(color=Color(line_color))

                    task_prefix = ''
                    task_sufix = line.project_activity_work_package_id.sap_report_task_sufix if line.project_activity_work_package_id.sap_report_task_sufix else ''
                    if project_wp_line and len(project_wp_line) == 1:
                        task_prefix = project_wp_line[0].sap_report_task_prefix if project_wp_line[0].sap_report_task_prefix else ''
                        task_sufix = project_wp_line[0].sap_report_task_sufix if project_wp_line[0].sap_report_task_sufix else ''
                    ws['I'+str(n)] = task_prefix + line.task_id.name + task_sufix if line.task_id else ''
                    ws['I'+str(n)].font = Font(color=Color(line_color))

                    ws['J'+str(n)] = ''
                    ws['J'+str(n)].font = Font(color=Color(line_color))
                    ws['K'+str(n)] = ''
                    ws['K'+str(n)].font = Font(color=Color(line_color))
                    ws['L'+str(n)] = line.account_id.name if line.account_id else ''
                    ws['L'+str(n)].font = Font(color=Color(line_color))
                    ws['M'+str(n)] = line.project_activity_work_package_id.name if line.project_activity_work_package_id else ''
                    ws['M'+str(n)].font = Font(color=Color(line_color))
                    ws['N'+str(n)] = line.project_activity_id.name if line.project_activity_id else ''
                    ws['N'+str(n)].font = Font(color=Color(line_color))
                    ws['O'+str(n)] = line.user_id.name
                    ws['O'+str(n)].font = Font(color=Color(line_color))
                    ws['P'+str(n)] = line.timesheet_comment if line.timesheet_comment else ''
                    ws['P'+str(n)].font = Font(color=Color(line_color))
                    ws['Q'+str(n)] = d.datetime.strptime(line.create_date, tools.DEFAULT_SERVER_DATETIME_FORMAT)
                    ws['Q'+str(n)].style = Style(alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'),
                                                 number_format="DD.MM.YYYY")
                    ws['Q'+str(n)].font = Font(color=Color(line_color))
                    line_posted_on_sap = '---'
                    ws['R'+str(n)] = line_posted_on_sap
                    ws['R'+str(n)].font = Font(color=Color(line_color))

                # hr leaves records
                lv_current_date = d.datetime.strptime(sap_from, tools.DEFAULT_SERVER_DATE_FORMAT)
                lv_end_date = d.datetime.strptime(sap_to, tools.DEFAULT_SERVER_DATE_FORMAT)
                while lv_current_date <= lv_end_date:
                    lv_is_holiday = self.env['hr.day.off'].search([('date', '=', lv_current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT))], limit=1)
                    if lv_is_holiday:
                        lv_current_date += d.timedelta(days=1)
                        continue

                    lv_lines = self.env['hr.holidays'].search([('date_from', '<=', lv_current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)),
                                                               ('date_to', '>=', lv_current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)),
                                                               ('state', '=', 'validate')])
                    if lv_lines and len(lv_lines)>0:
                        lv_line_color = 'ff8d02'
                        for line in lv_lines:
                            n += 1
                            emp = line.employee_id

                            ws['A'+str(n)] = emp.other_id if emp and emp.other_id else '---'
                            ws['A'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['B'+str(n)] = lv_current_date
                            ws['B'+str(n)].style = Style(font=Font(color=Color(lv_line_color)), alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'),
                                                         number_format="DD.MM.YYYY")

                            ws['C'+str(n)] = format_float_time(8.0)
                            ws['C'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['D'+str(n)] = format_float_time(15.695)
                            ws['D'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['E'+str(n)] = format_float_time(7.695)
                            ws['E'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['F'+str(n)] = '---'
                            ws['F'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['G'+str(n)] = ''
                            ws['G'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['H'+str(n)] = ''
                            ws['H'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['I'+str(n)] = ''
                            ws['I'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['J'+str(n)] = ''
                            ws['J'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['K'+str(n)] = ''
                            ws['K'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['L'+str(n)] = line.holiday_status_id.name
                            ws['L'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['M'+str(n)] = ''
                            ws['M'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['N'+str(n)] = ''
                            ws['N'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['O'+str(n)] = emp.user_id.name if emp.user_id else 'No related user for: ' + emp.name
                            ws['O'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['P'+str(n)] = ''
                            ws['P'+str(n)].font = Font(color=Color(lv_line_color))

                            ws['Q'+str(n)] = d.datetime.strptime(line.create_date, tools.DEFAULT_SERVER_DATETIME_FORMAT)
                            ws['Q'+str(n)].style = Style(font=Font(color=Color(lv_line_color)),
                                                         alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'),
                                                         number_format="DD.MM.YYYY")
                            line_posted_on_sap = '---'
                            ws['R'+str(n)] = line_posted_on_sap
                    lv_current_date += d.timedelta(days=1)





        buf = cStringIO.StringIO()
        wb.save(buf)
        buf.seek(0)
        out = base64.encodestring(buf.read())
        buf.close()

        file_name = 'employee_timesheets_'+dict(MONTHS).get(month)+'_'+dict(YEARS).get(year)
        if this.display_sap_report is True:
            week_no = datetime.strptime(this.sap_date_from, tools.DEFAULT_SERVER_DATE_FORMAT).isocalendar()[1]
            week = str(week_no) if week_no > 9 else '0' + str(week_no)
            file_name += '_W' + week

        self.write({'state': 'get',
                                  'data': out,
                                  'name': file_name + '.xlsx'})
        return {
            'type': 'ir.actions.act_window',
            'res_id': self[0].id,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'employee.timesheet.generator',
            'target': 'new',
            'context': self.env.context,
        }


def write_header(ws, employee, month, year):
    ws.merge_cells('A2:C2')
    ws['A2'] = 'TIMESHEET'
    ws['B2'].style = Style(border=Border(top=Side(style='thick', color=colors.BLACK)))
    ws['C2'].style = Style(border=Border(right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK)))

    ws['A2'].style = Style(Font(bold=True),
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('B3:C3')
    ws['A3'] = 'Employee:'
    ws['B3'] = employee.user_id.name if employee.user_id else 'No related user for: ' + employee.name
    ws['C3'].style = Style(border=Border(right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),))
    ws['B3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('B4:C4')
    ws['A4'] = 'ID number:'
    ws['B4'] = employee.other_id
    ws['C4'].style = Style(border=Border(right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),))
    ws['B4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('B5:C5')
    ws['A5'] = 'Month/Year:'
    ws['B5'] = dict(MONTHS).get(month)+' '+dict(YEARS).get(year)
    ws['B5'].font = Font(bold=True)
    ws['C5'].style = Style(border=Border(right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK),))
    ws['B5'].style = Style(Font(bold=True),
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws['A3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['A4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['A5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )

    ws.merge_cells('E2:G2')
    ws['E2'] = 'Holidays current month:'
    ws['H2'] = '---'
    ws['E2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['F2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['G2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['H2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('E3:G3')
    ws['E3'] = 'Holidays previous months:'
    ws['H3'] = '---'
    ws['E3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['F3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['G3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['H3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('E4:G4')
    ws['E4'] = 'Holidays credit on 31.12:'
    ws['H4'] = '---'
    ws['E4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['F4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['G4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['H4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('E5:G5')
    ws['E5'] = 'Holidays credit actual:'
    ws['H5'] = '---'
    ws['E5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['F5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['G5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['H5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('I2:K2')
    ws['I2'] = 'Sick leave current month:'
    ws['L2'] = '---'
    ws['I2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['J2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['K2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['L2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('I3:K3')
    ws['I3'] = 'Sick leave previous months:'
    ws['L3'] = '---'
    ws['I3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['J3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['K3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['L3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('I4:K4')
    ws['I4'] = 'Sick leave total:'
    ws['L4'] = '---'
    ws['I4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['J4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['K4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['L4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('M2:O2')
    ws['M2'] = 'Working hours actual value:'
    ws['P2'] = '---'
    ws['M2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['N2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['O2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['P2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('M3:O3')
    ws['M3'] = 'Working hours given value:'
    ws['P3'] = '---'
    ws['M3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['N3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['O3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['P3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('M4:O4')
    ws['M4'] = 'Balance of actual/given working hours:'
    ws['P4'] = '---'
    ws['M4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['N4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['O4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['P4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('M5:O5')
    ws['M5'] = 'Project time:'
    ws['P5'] = '---'
    ws['M5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['N5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['O5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['P5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('Q2:R2')
    ws['Q2'] = 'Overtime 125%:'
    ws['S2'] = '---'
    ws['Q2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['R2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['S2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('Q3:R3')
    ws['Q3'] = 'Overtime 150%:'
    ws['S3'] = '---'
    ws['Q3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['R3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['S3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('Q4:R4')
    ws['Q4'] = 'Overtime 200%:'
    ws['S4'] = '---'
    ws['Q4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['R4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['S4'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('Q5:R5')
    ws['Q5'] = 'TimeCredit utilised:'
    ws['S5'] = '---'
    ws['Q5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['R5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['S5'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('T2:U2')
    ws['T2'] = 'Overtime covered/Year:'
    ws['V2'] = '---'
    ws['T2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['U2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['V2'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    ws.merge_cells('T3:U3')
    ws['T3'] = 'Overtime actual/Year:'
    ws['V3'] = '---'
    ws['T3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['U3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           )
    ws['V3'].style = Style(
                           border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(horizontal='center')
                           )

    header_font = Font(bold=True)

    ws['A7'] = 'Date'
    ws['A7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['B7'] = 'Start'
    ws['B7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['C7'] = 'End'
    ws['C7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['D7'] = 'Break'
    ws['D7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['D7'].font = header_font
    ws['D7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['E7'] = 'Sum'
    ws['E7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['E7'].font = header_font
    ws['E7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['F7'] = 'Working time per day'
    ws['F7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['F7'].font = header_font
    ws['F7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['G7'] = 'Project'
    ws['G7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['G7'].font = header_font
    ws['G7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['H7'] = 'Activity'
    ws['H7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['H7'].font = header_font
    ws['H7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['I7'] = 'Site number'
    ws['I7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['I7'].font = header_font
    ws['I7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['J7'] = 'Start travel'
    ws['J7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['J7'].font = header_font
    ws['J7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['K7'] = 'End travel'
    ws['K7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['K7'].font = header_font
    ws['K7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['L7'] = 'Vehicle'
    ws['L7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['L7'].font = header_font
    ws['L7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['M7'] = 'Accommodation'
    ws['M7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['M7'].font = header_font
    ws['M7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['N7'] = 'Taggeld total'
    ws['N7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['N7'].font = header_font
    ws['N7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['O7'] = 'Taggeld taxable'
    ws['O7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['O7'].font = header_font
    ws['O7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['P7'] = 'Taggeld taxfree'
    ws['P7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['P7'].font = header_font
    ws['P7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))

    ws['Q7'] = 'Comment'
    ws['Q7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['Q7'].font = header_font
    ws['Q7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))
    ws['R7'] = 'Overtime 125%'
    ws['R7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['R7'].font = header_font
    ws['R7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))
    ws['S7'] = 'Overtime 150%'
    ws['S7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['S7'].font = header_font
    ws['S7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))
    ws['T7'] = 'Overtime 200%'
    ws['T7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['T7'].font = header_font
    ws['T7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))
    ws['U7'] = 'TimeCredit'
    ws['U7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['U7'].font = header_font
    ws['U7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))
    ws['V7'] = 'Accom. Recharge'
    ws['V7'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af'))
    ws['V7'].font = header_font
    ws['V7'].style = Style(border=Border(left=Side(style='thick', color=colors.BLACK),
                                         right=Side(style='thick', color=colors.BLACK),
                                         top=Side(style='thick', color=colors.BLACK),
                                         bottom=Side(style='thick', color=colors.BLACK)),
                           alignment=Alignment(wrap_text=True),
                           font=Font(bold=True),
                           fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ffe8af')))



    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[7].height = 45
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['Q'].width = 30


def write_line(ws, n, color, current_date, line=None):
    empty_line = False
    if line is None:
        empty_line = True

    ws['A'+str(7+n[0])] = print_date(line.date) if not empty_line and line.date else print_date(current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT))
    ws['A'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(horizontal='center',
                                                          wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['B'+str(7+n[0])] = format_float_time(line.timesheet_start_time) if not empty_line and line.timesheet_start_time else ''
    ws['B'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['C'+str(7+n[0])] = format_float_time(line.timesheet_end_time) if not empty_line and line.timesheet_end_time else ''
    ws['C'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['D'+str(7+n[0])] = format_float_time(line.timesheet_break_amount) if not empty_line and line.timesheet_break_amount else ''
    ws['D'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['E'+str(7+n[0])] = format_float_time(line.unit_amount) if not empty_line and line.unit_amount else ''
    ws['E'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['F'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['G'+str(7+n[0])] = line.account_id.name if not empty_line and line.account_id else ''
    ws['G'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['H'+str(7+n[0])] = line.project_activity_id.name if not empty_line and line.project_activity_id else ''
    ws['H'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['I'+str(7+n[0])] = line.task_id.name + '-1' if not empty_line and line.task_id else ''
    ws['I'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['J'+str(7+n[0])] = line.timesheet_travel_start if not empty_line and line.timesheet_travel_start else ''
    ws['J'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['K'+str(7+n[0])] = line.timesheet_travel_end if not empty_line and line.timesheet_travel_end else ''
    ws['K'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['L'+str(7+n[0])] = line.timesheet_vehicle_id.license_plate if not empty_line and line.timesheet_vehicle_id else ''
    ws['L'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                   )

    ws['M'+str(7+n[0])] = line.timesheet_accommodation if not empty_line and line.timesheet_accommodation else ''
    ws['M'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['N'+str(7+n[0])] = 0.0
    ws['N'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['O'+str(7+n[0])] = 0.0
    ws['O'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['P'+str(7+n[0])] = 0.0
    ws['P'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['Q'+str(7+n[0])] = line.timesheet_comment if not empty_line and line.timesheet_comment else ''
    ws['Q'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )
    ws['R'+str(7+n[0])] = 0.0
    ws['R'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['S'+str(7+n[0])] = 0.0
    ws['S'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['T'+str(7+n[0])] = 0.0
    ws['T'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['U'+str(7+n[0])] = 0.0
    ws['U'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['V'+str(7+n[0])] = 0.0
    ws['V'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )


def write_pub_holiday_line(ws, n, color, current_date):

    ws['A'+str(7+n[0])] = print_date(current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT))
    ws['A'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(horizontal='center',
                                                          wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['B'+str(7+n[0])] = format_float_time(8.0)
    ws['B'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['C'+str(7+n[0])] = format_float_time(15.695)
    ws['C'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['D'+str(7+n[0])] = format_float_time(0.0)
    ws['D'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['E'+str(7+n[0])] = format_float_time(7.695)
    ws['E'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['F'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['G'+str(7+n[0])] = 'Public Holiday'
    ws['G'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['H'+str(7+n[0])] = ''
    ws['H'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['I'+str(7+n[0])] = ''
    ws['I'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['J'+str(7+n[0])] = ''
    ws['J'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['K'+str(7+n[0])] = ''
    ws['K'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['L'+str(7+n[0])] = ''
    ws['L'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                   )

    ws['M'+str(7+n[0])] = ''
    ws['M'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['N'+str(7+n[0])] = 0.0
    ws['N'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['O'+str(7+n[0])] = 0.0
    ws['O'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['P'+str(7+n[0])] = 0.0
    ws['P'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['Q'+str(7+n[0])] = ''
    ws['Q'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )
    ws['R'+str(7+n[0])] = 0.0
    ws['R'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['S'+str(7+n[0])] = 0.0
    ws['S'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['T'+str(7+n[0])] = 0.0
    ws['T'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['U'+str(7+n[0])] = 0.0
    ws['U'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['V'+str(7+n[0])] = 0.0
    ws['V'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )


def write_leave_request_line(ws, n, color, activity, start_time, end_time, sum_time, current_date):

    ws['A'+str(7+n[0])] = print_date(current_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT))
    ws['A'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(horizontal='center',
                                                          wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['B'+str(7+n[0])] = format_float_time(start_time)
    ws['B'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['C'+str(7+n[0])] = format_float_time(end_time)
    ws['C'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['D'+str(7+n[0])] = format_float_time(0.0)
    ws['D'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['E'+str(7+n[0])] = format_float_time(sum_time)
    ws['E'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format="HH:MM:SS"
                                      )

    ws['F'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['G'+str(7+n[0])] = activity
    ws['G'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['H'+str(7+n[0])] = ''
    ws['H'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['I'+str(7+n[0])] = ''
    ws['I'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['J'+str(7+n[0])] = ''
    ws['J'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['K'+str(7+n[0])] = ''
    ws['K'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['L'+str(7+n[0])] = ''
    ws['L'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                   )

    ws['M'+str(7+n[0])] = ''
    ws['M'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )

    ws['N'+str(7+n[0])] = 0.0
    ws['N'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['O'+str(7+n[0])] = 0.0
    ws['O'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['P'+str(7+n[0])] = 0.0
    ws['P'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['Q'+str(7+n[0])] = ''
    ws['Q'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      )
    ws['R'+str(7+n[0])] = 0.0
    ws['R'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['S'+str(7+n[0])] = 0.0
    ws['S'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['T'+str(7+n[0])] = 0.0
    ws['T'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['U'+str(7+n[0])] = 0.0
    ws['U'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['V'+str(7+n[0])] = 0.0
    ws['V'+str(7+n[0])].style = Style(alignment=Alignment(wrap_text=True),
                                      fill=PatternFill(patternType='solid',
                                                       fill_type='solid',
                                                       fgColor=Color(color)),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thin', color=colors.BLACK),
                                                    bottom=Side(style='thin', color=colors.BLACK)),
                                      number_format='0.00'
                                      )


def write_footer(ws, n, working_time_sum, working_time_per_day_sum):
    n[0] += 1
    ws['A'+str(7+n[0])] = 'SUM'
    ws['A'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(horizontal='center',
                                                          wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      )
    ws['B'+str(7+n[0])] = ''
    ws['B'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['C'+str(7+n[0])] = ''
    ws['C'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['D'+str(7+n[0])] = ''
    ws['D'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['E'+str(7+n[0])] = format_float_time_str(working_time_sum)
    ws['E'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      )

    ws['F'+str(7+n[0])] = format_float_time_str(working_time_per_day_sum)
    ws['F'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      )

    ws['G'+str(7+n[0])] = ''
    ws['G'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['H'+str(7+n[0])] = ''
    ws['H'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['I'+str(7+n[0])] = ''
    ws['I'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['J'+str(7+n[0])] = ''
    ws['J'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['K'+str(7+n[0])] = ''
    ws['K'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['L'+str(7+n[0])] = ''
    ws['L'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )
    ws['M'+str(7+n[0])] = ''
    ws['M'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )
    ws['N'+str(7+n[0])] = 0.0
    ws['N'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['O'+str(7+n[0])] = 0.0
    ws['O'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['P'+str(7+n[0])] = 0.0
    ws['P'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['Q'+str(7+n[0])] = ''
    ws['Q'+str(7+n[0])].style = Style(border=Border(bottom=Side(style='thick',
                                                                color=colors.BLACK)),
                                      )

    ws['R'+str(7+n[0])] = 0.0
    ws['R'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['S'+str(7+n[0])] = 0.0
    ws['S'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['T'+str(7+n[0])] = 0.0
    ws['T'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
    ws['U'+str(7+n[0])] = 0.0
    ws['U'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      number_format='0.00'
                                      )

    ws['V'+str(7+n[0])] = 0.0
    ws['V'+str(7+n[0])].style = Style(font=Font(bold=True),
                                      alignment=Alignment(wrap_text=True),
                                      border=Border(left=Side(style='thick', color=colors.BLACK),
                                                    right=Side(style='thick', color=colors.BLACK),
                                                    top=Side(style='thick', color=colors.BLACK),
                                                    bottom=Side(style='thick', color=colors.BLACK)),
                                      number_format='0.00'
                                      )
