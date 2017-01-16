# -*- coding: utf-8 -*-

import openpyxl
import math
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, Style
import datetime
import base64
import cStringIO
from openerp import models, fields, api, tools, exceptions as e
from babel.dates import format_datetime


def print_date(date):
    day = datetime.datetime.strptime(date, tools.DEFAULT_SERVER_DATE_FORMAT)
    # res = format_datetime(day, format="d-MMM-yy", locale="en")
    return day

class ExportMilestonesWizard(models.TransientModel):
    _name = 'project.export.milestones.wizard'

    @api.onchange('project_id')
    def _change_project(self):
        if self.project_id:
            log_rec = self.env['project.export.milestones.log'].search([('project_id', '=', self.project_id.id)],
                                                                       limit=1,
                                                                       order='timestamp DESC')
            if log_rec:
                self.timestamp = log_rec.timestamp
                return
        res = datetime.datetime.now()
        res -= datetime.timedelta(days=1)
        self.timestamp = res.strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT)

    timestamp = fields.Datetime('Timestamp', required=True)
    project_id = fields.Many2one('project.project', 'Project')
    milestone_ids = fields.Many2many('project.milestone', string='Milestones', domain="[('project_id', '=', project_id)]")
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    report_type = fields.Selection([('fc', 'Forecasts'), ('ac', 'Actuals'), ('both', 'Both')], 'Report type', default='fc', required=True)
    state = fields.Selection([('choose', 'Choose'), ('get', 'Get')], 'State', default='choose')

    @api.multi
    def get_report(self):
        this = self[0]
        is_ac = False
        is_fc = False

        date_style = Style(number_format="DD.MM.YYYY")
        number_style = Style(number_format="0")
        date_limit = (datetime.datetime.now() - datetime.timedelta(days=14)).strftime(tools.DEFAULT_SERVER_DATE_FORMAT)

        if not this.project_id.project_code or not this.project_id.project_wp_code:
            raise e.UserError('Project ID and DWP ID are missing for ' + this.project_id.name + '.')

        wb = openpyxl.Workbook(encoding='utf-8')

        if len(this.milestone_ids) > 0:
            selected_milestones = this.milestone_ids
        else:
            selected_milestones = self.env['project.milestone'].search([('project_id', '=', this.project_id.id)], order='sequence DESC')

        for i in ['fc', 'ac']:

            if i == 'ac':
                if this.report_type not in ['ac', 'both']:
                    continue
                is_fc = False
                is_ac = True

            if i == 'fc':
                if this.report_type not in ['fc', 'both']:
                    continue
                is_fc = True
                is_ac = False

            for milestone in selected_milestones:

                # SERVER VERSION
                if is_fc is True:
                    ws = wb.create_sheet(0, milestone.name+'-FC')
                elif is_ac is True:
                    ws = wb.create_sheet(0, milestone.name+'-AC')

                # LOCAL VERSION
                # if is_fc is True:
                #     ws = wb.create_sheet(milestone.name+'-FC', 0)
                # elif is_ac is True:
                #     ws = wb.create_sheet(milestone.name+'-AC', 0)

                ws['A1'] = 'Project ID'
                ws['A2'] = int(this.project_id.project_code)
                # ws['A2'].style = number_style

                ws['B1'] = 'DWP ID'
                ws['B2'] = int(this.project_id.project_wp_code)
                # ws['B2'].style = number_style

                ws['C1'] = 'Site ID'

                ws['D1'] = 'WP ID'

                ws['E1'] = int(milestone.name)
                # ws['E1'].style = number_style

                if is_fc is True:
                    ws['E2'] = 'Forecast'
                    self.env.cr.execute("""
                        select
                            milestone_line_id
                        from
                            project_task_milestone_update_log
                        where
                            timestamp >= '%s'
                            and updated_field = 'forecast'
                    """ % this.timestamp)
                    logs = self.env.cr.fetchall()
                    line_ids = [log[0] for log in logs]

                    lines = self.env['project.task.milestone.forecast'].search([('project_id', '=', this.project_id.id),
                                                                                ('milestone_id', '=', milestone.id),
                                                                                ('forecast_date', '!=', False),
                                                                                ('forecast_date', '>=', date_limit),
                                                                                ('id', 'in', line_ids)])

                elif is_ac is True:
                    ws['E2'] = 'Actual'
                    self.env.cr.execute("""
                        select
                            milestone_line_id
                        from
                            project_task_milestone_update_log
                        where
                            timestamp >= '%s'
                            and updated_field = 'actual'
                    """ % this.timestamp)
                    logs = self.env.cr.fetchall()
                    line_ids = [log[0] for log in logs]
                    lines = self.env['project.task.milestone.forecast'].search([('project_id', '=', this.project_id.id),
                                                                                ('milestone_id', '=', milestone.id),
                                                                                ('actual_date', '!=', False),
                                                                                ('actual_date', '>=', date_limit),
                                                                                ('id', 'in', line_ids)])

                n = 2
                for line in lines:
                    n += 1
                    ws['C'+str(n)] = line.task_id.name

                    field = line.milestone_id.export_task_wp_code
                    wp_code = False
                    if field == 'sa':
                        wp_code = int(line.task_id.sa_work_package_code)
                    elif field == 'cw':
                        wp_code = int(line.task_id.cw_work_package_code)
                    elif field == 'ti':
                        wp_code = int(line.task_id.ti_work_package_code)
                    if not wp_code:
                        wp_code = '--- WP ID is missing ---'

                    ws['D'+str(n)] = wp_code
                    if is_fc is True:
                        ws['E'+str(n)] = print_date(line.forecast_date) if line.forecast_date else ''
                        ws['E'+str(n)].style = Style(number_format="DD.MM.YYYY")
                    elif is_ac is True:
                        ws['E'+str(n)] = print_date(line.actual_date) if line.actual_date else ''
                        if line.forecast_date == False:
                            ws['E'+str(n)].style = Style(number_format="DD.MM.YYYY")
                        else:
                            ws['E'+str(n)].style = Style(number_format="DD.MM.YYYY",
                                                         fill=PatternFill(patternType='solid',
                                                                          fill_type='solid',
                                                                          fgColor=Color('999a9e')))


        buf = cStringIO.StringIO()
        wb.save(buf)
        buf.seek(0)
        out = base64.encodestring(buf.read())
        buf.close()

        self.write({'state': 'get',
                    'data': out,
                    'name': 'mass_upload_data_'+datetime.datetime.now().strftime('%Y_%m_%d')+'.xlsx'})

        self.env['project.export.milestones.log'].create({
            'timestamp': datetime.datetime.now().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT),
            'project_id': this.project_id.id,
        })

        return {
            'type': 'ir.actions.act_window',
            'res_id': this.id,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'project.export.milestones.wizard',
            'target': 'new',
            'context': self.env.context,
        }
